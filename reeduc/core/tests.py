"""Tests for the core app."""

from datetime import date, datetime

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from django.contrib.auth import get_user_model

from .forms import CadastroForm, FamiliarForm
from .models import Agendamento, Atendimento, Cadastro, Familiar
from .views import _age_distribution, _report_querysets


class LoginPageTests(TestCase):
    """Smoke tests for the login page."""

    def test_login_page_renders(self):
        """Ensure the login page returns HTTP 200."""
        response = self.client.get(reverse("login"))
        self.assertEqual(response.status_code, 200)


class ReportFilterTests(TestCase):
    def setUp(self):
        self.matching = Cadastro.objects.create(
            nome="Maria da Silva", data_cadastro=date(2026, 6, 20), status="ativo"
        )
        self.other = Cadastro.objects.create(
            nome="Outra Pessoa", data_cadastro=date(2026, 5, 10), status="arquivado"
        )
        self.matching_relative = Familiar.objects.create(
            nome="Familiar Maria", cadastro=self.matching
        )
        self.other_relative = Familiar.objects.create(
            nome="Familiar Outra", cadastro=self.other
        )
        self.unlinked_relative = Familiar.objects.create(
            nome="Familiar Avulso", nome_interno_referencia="Egresso Referenciado"
        )

        january = timezone.make_aware(datetime(2026, 1, 15, 12))
        may = timezone.make_aware(datetime(2026, 5, 15, 12))
        Familiar.objects.filter(
            pk__in=[self.matching_relative.pk, self.unlinked_relative.pk]
        ).update(data_criacao=january)
        Familiar.objects.filter(pk=self.other_relative.pk).update(data_criacao=may)

    def test_family_period_uses_its_own_creation_date_and_keeps_unlinked(self):
        _, familiares, _, _ = _report_querysets({
            "data_inicio": "2026-01-01", "data_fim": "2026-01-31"
        })

        self.assertQuerySetEqual(
            familiares,
            [self.unlinked_relative, self.matching_relative],
            ordered=False,
        )

    def test_family_name_searches_familiar_and_referenced_egresso(self):
        _, familiares, _, _ = _report_querysets({"nome": "Referenciado"})

        self.assertQuerySetEqual(familiares, [self.unlinked_relative])

    def test_unfiltered_report_keeps_all_relatives(self):
        _, familiares, _, _ = _report_querysets({})

        self.assertEqual(familiares.count(), 3)

    def test_date_and_name_filter_activity_sections(self):
        matching_atendimento = Atendimento.objects.create(
            nome_pessoa_atendida="Maria da Silva", data_atendimento=date(2026, 6, 25)
        )
        Atendimento.objects.create(
            nome_pessoa_atendida="Maria da Silva", data_atendimento=date(2026, 5, 25)
        )
        matching_agendamento = Agendamento.objects.create(
            nome_atendido="Maria da Silva", data_agendamento=date(2026, 6, 26)
        )
        Agendamento.objects.create(
            nome_atendido="Outra Pessoa", data_agendamento=date(2026, 6, 26)
        )

        _, _, atendimentos, agendamentos = _report_querysets({
            "data_inicio": "2026-06-01",
            "data_fim": "2026-06-30",
            "nome": "Maria",
        })

        self.assertQuerySetEqual(atendimentos, [matching_atendimento])
        self.assertQuerySetEqual(agendamentos, [matching_agendamento])


class DicapFieldTests(TestCase):
    def test_cadastro_form_persists_multiple_deficiencies(self):
        form = CadastroForm(data={
            "nome": "Pessoa Egressa",
            "data_cadastro": "2026-07-19",
            "experiencia_trabalho": "formal_informal",
            "tipo_ocupacao": "formal",
            "grau_instrucao": "fundamental_incompleto",
            "serie_concluida": "5ª série",
            "deficiencias": ["visual", "auditiva"],
            "possui_comorbidade": "sim",
            "comorbidades": "Hipertensão",
            "uso_substancias_psicoativas": "uso_anterior",
            "substancias_psicoativas": "Álcool",
        })

        self.assertTrue(form.is_valid(), form.errors)
        cadastro = form.save()
        self.assertEqual(cadastro.deficiencias, "visual, auditiva")
        self.assertEqual(
            cadastro.get_deficiencias_display(),
            "Deficiência visual, Deficiência auditiva",
        )

    def test_familiar_form_persists_dicap_profile(self):
        cadastro = Cadastro.objects.create(nome="Egresso")
        form = FamiliarForm(
            data={
                "nome": "Familiar",
                "identidade_genero": "mulher_cis",
                "experiencia_trabalho": "informal",
                "ocupacao": "Costureira",
                "tipo_ocupacao": "informal",
                "grau_instrucao": "fundamental_incompleto",
                "serie_concluida": "3ª série",
                "deficiencias": ["motora"],
                "possui_comorbidade": "nao",
                "uso_substancias_psicoativas": "nunca",
            },
            instance=Familiar(cadastro=cadastro),
        )

        self.assertTrue(form.is_valid(), form.errors)
        familiar = form.save()
        self.assertEqual(familiar.ocupacao, "Costureira")
        self.assertEqual(familiar.serie_concluida, "3ª série")
        self.assertEqual(familiar.deficiencias, "motora")

    def test_official_age_bands_are_calculated(self):
        Cadastro.objects.create(nome="Jovem", data_nascimento=date(2002, 8, 1))
        Cadastro.objects.create(nome="Adulto", data_nascimento=date(1980, 1, 1))
        rows = dict(_age_distribution(Cadastro.objects.all()))

        self.assertEqual(rows["18 a 24"], 1)
        self.assertEqual(rows["35 a 59"], 1)

    def test_report_filters_new_fields_for_cadastros_and_familiares(self):
        cadastro = Cadastro.objects.create(
            nome="Com perfil",
            experiencia_trabalho="formal",
            tipo_ocupacao="formal",
            possui_comorbidade="sim",
        )
        Familiar.objects.create(
            nome="Familiar com perfil",
            cadastro=cadastro,
            experiencia_trabalho="formal",
            tipo_ocupacao="formal",
            possui_comorbidade="sim",
        )
        Cadastro.objects.create(nome="Sem perfil")
        Familiar.objects.create(nome="Outro familiar")

        cadastros, familiares, _, _ = _report_querysets({
            "experiencia_trabalho": "formal",
            "tipo_ocupacao": "formal",
            "possui_comorbidade": "sim",
        })

        self.assertEqual(cadastros.count(), 1)
        self.assertEqual(familiares.count(), 1)

    def test_pdf_report_with_dicap_sections_is_generated(self):
        user = get_user_model().objects.create(username="relatorio")
        Cadastro.objects.create(
            nome="Pessoa & Teste",
            data_nascimento=date(1990, 1, 1),
            experiencia_trabalho="formal",
            tipo_ocupacao="formal",
            grau_instrucao="medio_completo",
            serie_concluida="3ª série",
            deficiencias="visual",
            possui_comorbidade="sim",
            comorbidades="Hipertensão",
            uso_substancias_psicoativas="uso_anterior",
        )
        self.client.force_login(user)

        response = self.client.get(reverse("relatorios-pdf"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertGreater(len(response.content), 1000)

    def test_excel_report_is_formatted_and_contains_extractable_sheets(self):
        from io import BytesIO

        from openpyxl import load_workbook

        user = get_user_model().objects.create(username="excel")
        cadastro = Cadastro.objects.create(
            nome="Pessoa para Excel",
            data_nascimento=date(1990, 1, 1),
            experiencia_trabalho="formal",
            tipo_ocupacao="formal",
            grau_instrucao="medio_completo",
            serie_concluida="3ª série",
            deficiencias="visual",
        )
        Familiar.objects.create(
            nome="Familiar para Excel",
            cadastro=cadastro,
            ocupacao="Costureira",
            grau_instrucao="fundamental_incompleto",
            serie_concluida="5ª série",
        )
        self.client.force_login(user)

        response = self.client.get(reverse("relatorios-excel"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(
            workbook.sheetnames,
            ["Resumo", "Egressos", "Familiares", "Atendimentos", "Agendamentos"],
        )
        egressos = workbook["Egressos"]
        self.assertEqual(egressos.freeze_panes, "A4")
        self.assertTrue(egressos.auto_filter.ref)
        self.assertEqual(len(egressos.tables), 1)
        self.assertEqual(egressos["B4"].value, "Pessoa para Excel")
        self.assertIn("Série final", [cell.value for cell in egressos[3]])
        familiares = workbook["Familiares"]
        self.assertEqual(familiares["B4"].value, "Familiar para Excel")


class PremiumRegistrationFormTests(TestCase):
    def test_global_notifications_render_success_messages(self):
        user = get_user_model().objects.create(username="notificacoes")
        atendimento = Atendimento.objects.create(
            nome_pessoa_atendida="Pessoa",
            data_atendimento=date(2026, 7, 19),
            status="pendente",
        )
        self.client.force_login(user)

        response = self.client.post(
            reverse("atendimento-realizar", args=[atendimento.pk]),
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Atendimento marcado como realizado.")
        self.assertContains(response, 'id="appToastRegion"')
        self.assertContains(response, "notifications.css")
        atendimento.refresh_from_db()
        self.assertEqual(atendimento.status, "realizado")

    def test_choice_fields_use_clear_prompts(self):
        cadastro_form = CadastroForm()
        familiar_form = FamiliarForm(instance=Familiar(cadastro=Cadastro(nome="Egresso")))

        self.assertEqual(
            list(cadastro_form.fields["identidade_genero"].choices)[0],
            ("", "Selecione uma opção"),
        )
        self.assertEqual(
            list(familiar_form.fields["tipo_ocupacao"].choices)[0],
            ("", "Selecione uma opção"),
        )

    def test_hidden_conditional_details_are_cleared_by_server_validation(self):
        form = CadastroForm(data={
            "nome": "Pessoa sem condições declaradas",
            "data_cadastro": "2026-07-19",
            "possui_comorbidade": "nao",
            "comorbidades": "Valor antigo",
            "uso_substancias_psicoativas": "nunca",
            "substancias_psicoativas": "Valor antigo",
            "fez_ensino_superior": "nao",
            "curso_superior": "Valor antigo",
            "estuda_atualmente": "nao",
            "horario_turno_estudo": "Valor antigo",
        })

        self.assertTrue(form.is_valid(), form.errors)
        cadastro = form.save()
        self.assertEqual(cadastro.comorbidades, "")
        self.assertEqual(cadastro.substancias_psicoativas, "")
        self.assertEqual(cadastro.curso_superior, "")
        self.assertEqual(cadastro.horario_turno_estudo, "")

    def test_main_registration_pages_render_premium_assets_and_required_fields(self):
        user = get_user_model().objects.create(username="formularios")
        self.client.force_login(user)

        cadastro_response = self.client.get(reverse("cadastro-adicionar"))
        familiar_response = self.client.get(reverse("familiar-avulso-adicionar"))

        self.assertContains(cadastro_response, "forms-premium.css")
        self.assertContains(cadastro_response, 'id="id_nome"')
        self.assertContains(cadastro_response, "required")
        self.assertContains(familiar_response, "Trabalho, escolaridade e saúde")
        self.assertContains(familiar_response, 'class="form-section"')

    def test_all_operational_registration_pages_render(self):
        user = get_user_model().objects.create(username="fluxos")
        self.client.force_login(user)
        cadastro = Cadastro.objects.create(nome="Pessoa")
        familiar = Familiar.objects.create(nome="Familiar", cadastro=cadastro)
        atendimento = Atendimento.objects.create(
            nome_pessoa_atendida="Pessoa", data_atendimento=date(2026, 7, 19)
        )
        agendamento = Agendamento.objects.create(
            nome_atendido="Pessoa", data_agendamento=date(2026, 7, 20)
        )

        urls = [
            reverse("cadastro-adicionar"),
            reverse("cadastro-editar", args=[cadastro.pk]),
            reverse("familiar-avulso-adicionar"),
            reverse("familiar-editar", args=[familiar.pk]),
            reverse("atendimento-adicionar"),
            reverse("atendimento-editar", args=[atendimento.pk]),
            reverse("agendamento-adicionar"),
            reverse("agendamento-editar", args=[agendamento.pk]),
        ]

        for url in urls:
            with self.subTest(url=url):
                response = self.client.get(url)
                self.assertEqual(response.status_code, 200)
                self.assertContains(response, "forms-premium.css")
