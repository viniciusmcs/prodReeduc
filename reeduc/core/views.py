"""Core views.

Keep views thin: business logic stays minimal and delegated.
"""

from datetime import datetime

from django.contrib.auth import authenticate, get_user_model, login, logout
from django.contrib import messages
from django.conf import settings
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import IntegrityError, models as db_models
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from .forms import AdminUserCreateForm, AgendamentoForm, AtendimentoForm, CadastroForm, FamiliarForm, LembreteForm

from .models import Agendamento, Atendimento, Cadastro, Familiar, Lembrete, UserProfile

def home_redirect(request):
    """Redirect root URL to the login screen with redirect parameter."""
    if request.user.is_authenticated:
        return redirect("/home/")
    return redirect("/autenticar/entrar?redirect=%2Fhome%2F")


@login_required
def home_view(request):
    """Render the home dashboard layout."""
    agendamentos = Agendamento.objects.all().order_by("-data_agendamento")[:10]
    
    # Add cadastro object to each agendamento for linking
    agendamentos_with_cadastro = []
    for agendamento in agendamentos:
        cadastro = Cadastro.objects.filter(nome=agendamento.nome_atendido).first()
        agendamentos_with_cadastro.append({
            'agendamento': agendamento,
            'cadastro': cadastro
        })
    
    # Get cadastro statistics
    total_cadastros = Cadastro.objects.count()
    cadastros_ativos = Cadastro.objects.filter(status="ativo").count()
    cadastros_arquivados = Cadastro.objects.filter(status="arquivado").count()
    ultimo_cadastro = Cadastro.objects.order_by("-id").first()
    
    # Get other counts
    total_atendimentos = Atendimento.objects.count()
    total_familiares_avulsos = Familiar.objects.filter(cadastro__isnull=True).count()
    
    context = {
        "user_name": request.user.get_username(),
        "agendamentos_with_cadastro": agendamentos_with_cadastro,
        "total_agendamentos": Agendamento.objects.count(),
        "total_cadastros": total_cadastros,
        "cadastros_ativos": cadastros_ativos,
        "cadastros_arquivados": cadastros_arquivados,
        "ultimo_cadastro": ultimo_cadastro,
        "total_atendimentos": total_atendimentos,
        "total_familiares_avulsos": total_familiares_avulsos,
    }
    return render(request, "core/home.html", context)


@login_required
def cadastro_adicionar_view(request):
    """Create a new cadastro with the requested identification fields."""
    if request.method == "POST":
        form = CadastroForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Cadastro realizado com sucesso.")
            return redirect("cadastro-lista")
    else:
        form = CadastroForm()

    context = {
        "form": form,
    }
    return render(request, "core/cadastro_adicionar.html", context)


@login_required
def cadastro_lista_view(request, filtro: str | None = None):
    """List cadastros based on filter (todos/ativos/arquivados/familiares)."""
    cadastros = Cadastro.objects.all().order_by("-data_cadastro")
    filtro_label = "Todos"

    if filtro == "ativos":
        cadastros = cadastros.filter(status="ativo")
        filtro_label = "Ativos"
    elif filtro == "arquivados":
        cadastros = cadastros.filter(status="arquivado")
        filtro_label = "Arquivados"
    elif filtro == "familiares":
        return redirect("familiares-avulsos-lista")

    context = {
        "cadastros": cadastros,
        "filtro": filtro_label,
    }
    return render(request, "core/cadastro_lista.html", context)


@login_required
def cadastro_dashboard_view(request):
    """Render the cadastro dashboard with search and status lists."""
    query = request.GET.get("q", "").strip()
    cadastros = Cadastro.objects.all().order_by("-data_cadastro")
    if query:
        cadastros = cadastros.filter(nome__icontains=query)

    ativos = cadastros.filter(status="ativo")
    arquivados = cadastros.filter(status="arquivado")

    context = {
        "query": query,
        "ativos": ativos,
        "arquivados": arquivados,
        "total": cadastros.count(),
    }
    return render(request, "core/cadastro_dashboard.html", context)


@login_required
def cadastro_perfil_view(request, cadastro_id: int):
    """Render the cadastro profile page with navigation tabs."""
    cadastro = get_object_or_404(Cadastro, id=cadastro_id)
    return render(request, "core/cadastro_perfil.html", {"cadastro": cadastro})


@login_required
def cadastro_familiares_view(request, cadastro_id: int):
    """Render and manage familiares for a cadastro."""
    cadastro = get_object_or_404(Cadastro, id=cadastro_id)
    if request.method == "POST" and request.POST.get("familiar_form") == "1":
        familiar_form = FamiliarForm(
            request.POST,
            request.FILES,
            instance=Familiar(cadastro=cadastro),
        )
        if familiar_form.is_valid():
            familiar_form.save()
            messages.success(request, "Familiar vinculado com sucesso.")
            return redirect(f"/cadastro/familiares/{cadastro_id}")
    else:
        familiar_form = FamiliarForm(instance=Familiar(cadastro=cadastro))

    familiares = cadastro.familiares.all().order_by("-data_criacao")
    return render(
        request,
        "core/cadastro_familiares.html",
        {
            "cadastro": cadastro,
            "familiares": familiares,
            "familiar_form": familiar_form,
        },
    )


@login_required
def cadastro_atendimentos_view(request, cadastro_id: int):
    """Render atendimentos history for a cadastro."""
    cadastro = get_object_or_404(Cadastro, id=cadastro_id)
    atendimentos = Atendimento.objects.filter(
        nome_pessoa_atendida__iexact=cadastro.nome
    ).order_by("-data_atendimento")
    return render(
        request,
        "core/cadastro_atendimentos.html",
        {"cadastro": cadastro, "atendimentos": atendimentos},
    )


@login_required
def cadastro_agendamentos_view(request, cadastro_id: int):
    """Render agendamentos history for a cadastro."""
    cadastro = get_object_or_404(Cadastro, id=cadastro_id)
    agendamentos = Agendamento.objects.filter(
        nome_atendido__iexact=cadastro.nome
    ).order_by("-data_agendamento")
    return render(
        request,
        "core/cadastro_agendamentos.html",
        {"cadastro": cadastro, "agendamentos": agendamentos},
    )


@login_required
def familiar_ver_view(request, familiar_id: int):
    """View familiar details."""
    familiar = get_object_or_404(Familiar, id=familiar_id)
    if familiar.cadastro_id is None:
        documentos_possui = [
            item.strip() for item in (familiar.documentos_possui or "").split(",") if item.strip()
        ]
        documentos_ausentes = [
            item.strip() for item in (familiar.documentos_ausentes or "").split(",") if item.strip()
        ]
        return render(
            request,
            "core/familiar_avulso_ver.html",
            {
                "familiar": familiar,
                "documentos_possui": documentos_possui,
                "documentos_ausentes": documentos_ausentes,
            },
        )

    documentos_possui = [
        item.strip() for item in (familiar.documentos_possui or "").split(",") if item.strip()
    ]
    documentos_ausentes = [
        item.strip() for item in (familiar.documentos_ausentes or "").split(",") if item.strip()
    ]
    return render(
        request,
        "core/familiar_ver.html",
        {
            "familiar": familiar,
            "documentos_possui": documentos_possui,
            "documentos_ausentes": documentos_ausentes,
        },
    )


@login_required
def familiar_upload_foto_view(request, familiar_id: int):
    """Handle photo upload for a familiar."""
    familiar = get_object_or_404(Familiar, id=familiar_id)

    if request.method == "POST" and request.POST.get("remover_foto") == "1":
        if familiar.foto:
            familiar.foto.delete(save=False)
        familiar.foto = None
        familiar.save()
        return JsonResponse(
            {"status": "success", "message": "Foto removida com sucesso", "foto_url": None}
        )

    if request.method == "POST" and request.FILES.get("foto"):
        foto = request.FILES["foto"]

        valid_extensions = [".png", ".jpeg", ".jpg"]
        file_ext = "." + foto.name.split(".")[-1].lower() if "." in foto.name else ""

        if file_ext not in valid_extensions:
            return JsonResponse(
                {"status": "error", "message": "Por favor, selecione apenas arquivos PNG ou JPEG"},
                status=400,
            )

        familiar.foto = foto
        familiar.save()

        return JsonResponse(
            {"status": "success", "message": "Foto enviada com sucesso", "foto_url": familiar.foto.url}
        )

    return JsonResponse({"status": "error", "message": "Requisição inválida"}, status=400)


@login_required
def familiar_editar_view(request, familiar_id: int):
    """Edit familiar details."""
    familiar = get_object_or_404(Familiar, id=familiar_id)
    if request.method == "POST":
        form = FamiliarForm(request.POST, request.FILES, instance=familiar)
        if form.is_valid():
            form.save()
            messages.success(request, "Dados do familiar atualizados com sucesso.")
            if familiar.cadastro_id is None:
                return redirect("familiares-avulsos-lista")
            return redirect(f"/cadastro/familiares/{familiar.cadastro_id}")
    else:
        form = FamiliarForm(instance=familiar)

    return render(request, "core/familiar_editar.html", {"form": form, "familiar": familiar})


@login_required
def familiar_excluir_view(request, familiar_id: int):
    """Delete familiar after confirmation."""
    familiar = get_object_or_404(Familiar, id=familiar_id)
    if request.method == "POST":
        cadastro_id = familiar.cadastro_id
        familiar.delete()
        messages.success(request, "Familiar excluído com sucesso.")
        if cadastro_id is None:
            return redirect("familiares-avulsos-lista")
        return redirect(f"/cadastro/familiares/{cadastro_id}")
    return render(request, "core/familiar_excluir.html", {"familiar": familiar})


@login_required
def familiar_avulso_adicionar_view(request):
    """Create a familiar record without vínculo com cadastro."""
    if request.method == "POST":
        form = FamiliarForm(request.POST, request.FILES)
        if form.is_valid():
            familiar = form.save(commit=False)
            familiar.cadastro = None
            familiar.save()
            messages.success(request, "Familiar cadastrado com sucesso.")
            return redirect("familiares-avulsos-lista")
    else:
        form = FamiliarForm()

    return render(request, "core/familiar_avulso_adicionar.html", {"form": form})


@login_required
def familiares_avulsos_lista_view(request):
    """List familiar records without vínculo com cadastro."""
    query = request.GET.get("q", "").strip()
    familiares = Familiar.objects.filter(cadastro__isnull=True).order_by("-data_criacao")
    if query:
        familiares = familiares.filter(nome__icontains=query)

    return render(
        request,
        "core/familiares_avulsos_lista.html",
        {
            "familiares": familiares,
            "query": query,
            "total": familiares.count(),
        },
    )


@login_required
def familiar_avulso_ver_view(request, familiar_id: int):
    """View a familiar record without vínculo com cadastro."""
    familiar = get_object_or_404(Familiar, id=familiar_id, cadastro__isnull=True)
    documentos_possui = [
        item.strip() for item in (familiar.documentos_possui or "").split(",") if item.strip()
    ]
    documentos_ausentes = [
        item.strip() for item in (familiar.documentos_ausentes or "").split(",") if item.strip()
    ]
    return render(
        request,
        "core/familiar_avulso_ver.html",
        {
            "familiar": familiar,
            "documentos_possui": documentos_possui,
            "documentos_ausentes": documentos_ausentes,
        },
    )


@login_required
def familiar_avulso_editar_view(request, familiar_id: int):
    """Edit a familiar record without vínculo com cadastro."""
    familiar = get_object_or_404(Familiar, id=familiar_id, cadastro__isnull=True)
    if request.method == "POST":
        form = FamiliarForm(request.POST, request.FILES, instance=familiar)
        if form.is_valid():
            familiar = form.save(commit=False)
            familiar.cadastro = None
            familiar.save()
            messages.success(request, "Dados do familiar atualizados com sucesso.")
            return redirect("familiares-avulsos-lista")
    else:
        form = FamiliarForm(instance=familiar)

    return render(
        request,
        "core/familiar_avulso_editar.html",
        {"form": form, "familiar": familiar},
    )


@login_required
def familiar_avulso_excluir_view(request, familiar_id: int):
    """Delete a familiar record without vínculo com cadastro."""
    familiar = get_object_or_404(Familiar, id=familiar_id, cadastro__isnull=True)
    if request.method == "POST":
        familiar.delete()
        messages.success(request, "Familiar excluído com sucesso.")
        return redirect("familiares-avulsos-lista")
    return render(request, "core/familiar_avulso_excluir.html", {"familiar": familiar})


@login_required
def cadastro_upload_foto_view(request, cadastro_id: int):
    """Handle photo upload for a cadastro."""
    from django.http import JsonResponse
    
    cadastro = get_object_or_404(Cadastro, id=cadastro_id)
    
    if request.method == "POST" and request.POST.get("remover_foto") == "1":
        if cadastro.foto:
            cadastro.foto.delete(save=False)
        cadastro.foto = None
        cadastro.save()
        return JsonResponse({
            "status": "success",
            "message": "Foto removida com sucesso",
            "foto_url": None,
        })

    if request.method == "POST" and request.FILES.get("foto"):
        foto = request.FILES["foto"]
        
        # Validar extensão
        valid_extensions = [".png", ".jpeg", ".jpg"]
        file_ext = "." + foto.name.split(".")[-1].lower() if "." in foto.name else ""
        
        if file_ext not in valid_extensions:
            return JsonResponse({
                "status": "error",
                "message": "Por favor, selecione apenas arquivos PNG ou JPEG"
            }, status=400)
        
        # Salvar foto
        cadastro.foto = foto
        cadastro.save()
        
        return JsonResponse({
            "status": "success",
            "message": "Foto enviada com sucesso",
            "foto_url": cadastro.foto.url
        })
    
    return JsonResponse({"status": "error", "message": "Requisição inválida"}, status=400)


@login_required
def anotacoes_editar_view(request):
    """Create or edit lembretes for a cadastro."""
    cadastro_id = request.GET.get("cadastro_id")
    cadastro = None
    lembrete = None

    if cadastro_id:
        cadastro = get_object_or_404(Cadastro, id=cadastro_id)
        lembrete = Lembrete.objects.filter(cadastro=cadastro).order_by("-data_atualizacao").first()

    if request.method == "POST":
        form = LembreteForm(request.POST, instance=lembrete)
        if form.is_valid():
            lembrete_obj = form.save(commit=False)
            if cadastro:
                lembrete_obj.cadastro = cadastro
            if not lembrete_obj.criado_por:
                lembrete_obj.criado_por = request.user
            lembrete_obj.atualizado_por = request.user
            lembrete_obj.save()
            messages.success(request, "Anotação salva com sucesso.")
            if cadastro:
                return redirect("cadastro-perfil", cadastro_id=cadastro.id)
            return redirect("home")
    else:
        form = LembreteForm(instance=lembrete)

    context = {
        "form": form,
        "cadastro": cadastro,
        "lembrete": lembrete,
    }
    return render(request, "core/anotacoes_editar.html", context)


@login_required
def cadastro_detalhe_view(request, cadastro_id: int):
    """Show details for a single cadastro."""
    cadastro = get_object_or_404(Cadastro, id=cadastro_id)
    return render(request, "core/cadastro_detalhe.html", {"cadastro": cadastro})


@login_required
def cadastro_editar_view(request, cadastro_id: int):
    """Edit an existing cadastro."""
    cadastro = get_object_or_404(Cadastro, id=cadastro_id)
    if request.method == "POST":
        form = CadastroForm(request.POST, request.FILES, instance=cadastro)
        if form.is_valid():
            obj = form.save(commit=False)
            if request.POST.get("remover_foto") == "1" and not request.FILES.get("foto"):
                obj.foto = None
            obj.save()
            messages.success(request, "Cadastro atualizado com sucesso.")
            return redirect("cadastro-lista")
    else:
        form = CadastroForm(instance=cadastro)

    return render(request, "core/cadastro_editar.html", {"form": form, "cadastro": cadastro})


@login_required
def cadastro_excluir_view(request, cadastro_id: int):
    """Delete a cadastro after confirmation."""
    cadastro = get_object_or_404(Cadastro, id=cadastro_id)
    if request.method == "POST":
        cadastro.delete()
        messages.success(request, "Cadastro excluído com sucesso.")
        return redirect("cadastro-lista")
    return render(request, "core/cadastro_excluir.html", {"cadastro": cadastro})


@login_required
def atendimento_adicionar_view(request):
    """Create a new atendimento record."""
    if request.method == "POST":
        form = AtendimentoForm(request.POST, user=request.user)
        if form.is_valid():
            atendimento = form.save(commit=False)
            atendimento.profissional_responsavel = request.user.get_username()
            atendimento.save()
            messages.success(request, "Atendimento registrado com sucesso.")
            return redirect("home")
    else:
        form = AtendimentoForm(
            user=request.user,
            initial={
                "profissional_responsavel": request.user.get_username(),
                "data_atendimento": timezone.localdate(),
            }
        )

    return render(request, "core/atendimento_adicionar.html", {"form": form})


@login_required
def atendimentos_dashboard_view(request):
    """List atendimentos with filters."""
    query = request.GET.get("q", "").strip()
    data = request.GET.get("data", "").strip()
    status = request.GET.get("status", "").strip()

    atendimentos = Atendimento.objects.all().order_by("-data_atendimento")
    if query:
        atendimentos = atendimentos.filter(nome_pessoa_atendida__icontains=query)
    if data:
        atendimentos = atendimentos.filter(data_atendimento=data)
    if status:
        atendimentos = atendimentos.filter(status=status)

    return render(
        request,
        "core/atendimentos_dashboard.html",
        {
            "atendimentos": atendimentos,
            "query": query,
            "data": data,
            "status": status,
            "total": atendimentos.count(),
        },
    )


@login_required
def atendimento_realizado_view(request, atendimento_id: int):
    """Mark atendimento as realizado."""
    atendimento = get_object_or_404(Atendimento, id=atendimento_id)
    if request.method == "POST":
        atendimento.status = "realizado"
        atendimento.save()
        messages.success(request, "Atendimento marcado como realizado.")
    return redirect(f"/atendimentos/{atendimento_id}/ver")


@login_required
def atendimento_ver_view(request, atendimento_id: int):
    """View atendimento details."""
    atendimento = get_object_or_404(Atendimento, id=atendimento_id)
    return render(request, "core/atendimento_ver.html", {"atendimento": atendimento})


@login_required
def atendimento_editar_view(request, atendimento_id: int):
    """Edit an existing atendimento."""
    atendimento = get_object_or_404(Atendimento, id=atendimento_id)
    if request.method == "POST":
        form = AtendimentoForm(request.POST, instance=atendimento)
        if form.is_valid():
            form.save()
            messages.success(request, "Atendimento atualizado com sucesso.")
            return redirect("home")
    else:
        form = AtendimentoForm(instance=atendimento)

    return render(request, "core/atendimento_editar.html", {"form": form, "atendimento": atendimento})


@login_required
def atendimento_excluir_view(request, atendimento_id: int):
    """Delete an atendimento after confirmation."""
    atendimento = get_object_or_404(Atendimento, id=atendimento_id)
    if request.method == "POST":
        atendimento.delete()
        messages.success(request, "Atendimento excluído com sucesso.")
        return redirect("home")
    return render(request, "core/atendimento_excluir.html", {"atendimento": atendimento})


@login_required
def agendamento_adicionar_view(request):
    """Create a new agendamento record."""
    if request.method == "POST":
        form = AgendamentoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Agendamento criado com sucesso.")
            return redirect("agendamentos-dashboard")
    else:
        form = AgendamentoForm(initial={"data_agendamento": timezone.localdate()})

    return render(request, "core/agendamento_adicionar.html", {"form": form})


@login_required
def agendamento_ver_view(request, agendamento_id: int):
    """View agendamento details."""
    agendamento = get_object_or_404(Agendamento, id=agendamento_id)
    return render(request, "core/agendamento_ver.html", {"agendamento": agendamento})


@login_required
def agendamento_editar_view(request, agendamento_id: int):
    """Edit an existing agendamento."""
    agendamento = get_object_or_404(Agendamento, id=agendamento_id)
    if request.method == "POST":
        form = AgendamentoForm(request.POST, instance=agendamento)
        if form.is_valid():
            form.save()
            messages.success(request, "Agendamento atualizado com sucesso.")
            return redirect("agendamentos-dashboard")
    else:
        form = AgendamentoForm(instance=agendamento)
    
    return render(request, "core/agendamento_editar.html", {"form": form, "agendamento": agendamento})


@login_required
def agendamento_excluir_view(request, agendamento_id: int):
    """Delete an agendamento after confirmation."""
    agendamento = get_object_or_404(Agendamento, id=agendamento_id)
    if request.method == "POST":
        agendamento.delete()
        messages.success(request, "Agendamento excluído com sucesso.")
        return redirect("agendamentos-dashboard")
    return render(request, "core/agendamento_excluir.html", {"agendamento": agendamento})


@login_required
def user_profile_view(request, username: str):
    """Render the user profile details screen."""
    if request.user.username != username and not request.user.is_superuser:
        return redirect("home")
    return render(
        request,
        "core/user_profile.html",
        {
            "profile_user": request.user,
        },
    )


@login_required
def user_profile_update_view(request, username: str):
    """Update the user's name or username."""
    if request.user.username != username and not request.user.is_superuser:
        return redirect("home")

    profile_update_error = None
    profile_update_success = None

    if request.method == "POST":
        username_value = request.POST.get("username", "").strip()
        full_name_value = request.POST.get("full_name", "").strip()

        if not username_value:
            profile_update_error = "Informe o nome de usuário."
        else:
            User = get_user_model()
            username_exists = (
                User.objects.filter(username__iexact=username_value)
                .exclude(id=request.user.id)
                .exists()
            )
            if username_exists:
                profile_update_error = "Nome de usuário já está em uso."

        if not profile_update_error:
            request.user.username = username_value
            if full_name_value:
                parts = full_name_value.split()
                request.user.first_name = parts[0]
                request.user.last_name = " ".join(parts[1:])
            else:
                request.user.first_name = ""
                request.user.last_name = ""
            request.user.save()
            messages.success(request, "Dados do perfil atualizados com sucesso.")
            return redirect(f"/usuario/perfil/{request.user.username}/editar")
        messages.error(request, profile_update_error)

    return render(
        request,
        "core/user_profile_update.html",
        {
            "profile_user": request.user,
            "profile_update_error": profile_update_error,
            "profile_update_success": profile_update_success,
        },
    )


@login_required
def user_profile_password_view(request, username: str):
    """Update the user's password."""
    if request.user.username != username and not request.user.is_superuser:
        return redirect("home")

    password_error = None
    password_success = None

    if request.method == "POST":
        current_password = request.POST.get("current_password", "")
        new_password = request.POST.get("new_password", "")
        confirm_password = request.POST.get("confirm_password", "")

        if not request.user.check_password(current_password):
            password_error = "Senha atual incorreta."
        elif not new_password:
            password_error = "Informe a nova senha."
        elif new_password != confirm_password:
            password_error = "As senhas não conferem."

        if not password_error:
            request.user.set_password(new_password)
            request.user.save()
            from django.contrib.auth import update_session_auth_hash

            update_session_auth_hash(request, request.user)
            messages.success(request, "Senha alterada com sucesso.")
        else:
            messages.error(request, password_error)

    return render(
        request,
        "core/user_profile_password.html",
        {
            "profile_user": request.user,
            "password_error": password_error,
            "password_success": password_success,
        },
    )


@login_required
def minhas_atividades_view(request):
    """Render the activities filter screen."""
    tipo = request.GET.get("tipo_atividade", "").strip()
    data_inicio_raw = request.GET.get("data_inicio", "").strip()
    data_fim_raw = request.GET.get("data_fim", "").strip()

    data_inicio = None
    data_fim = None
    if data_inicio_raw:
        data_inicio = datetime.strptime(data_inicio_raw, "%Y-%m-%d").date()
    if data_fim_raw:
        data_fim = datetime.strptime(data_fim_raw, "%Y-%m-%d").date()

    atividades = []

    def should_include(current_tipo: str) -> bool:
        return tipo in ("", "todos", current_tipo)

    def normalize_datetime(value):
        if timezone.is_naive(value):
            return timezone.make_aware(value, timezone.get_current_timezone())
        return timezone.localtime(value)

    if should_include("cadastros"):
        cadastros = Cadastro.objects.all()
        if data_inicio:
            cadastros = cadastros.filter(data_cadastro__gte=data_inicio)
        if data_fim:
            cadastros = cadastros.filter(data_cadastro__lte=data_fim)
        for cadastro in cadastros:
            atividades.append(
                {
                    "data": cadastro.data_cadastro,
                    "data_ordem": normalize_datetime(
                        datetime.combine(cadastro.data_cadastro, datetime.min.time())
                    ),
                    "atividade": "Novo Cadastro",
                    "tecnico": "Sistema",
                    "egresso": cadastro.nome,
                    "acoes": {
                        "ver": f"/home/cadastros/{cadastro.id}/ver",
                        "editar": f"/home/cadastros/{cadastro.id}/editar",
                        "excluir": f"/home/cadastros/{cadastro.id}/excluir",
                    },
                }
            )

    if should_include("agendamentos"):
        agendamentos = Agendamento.objects.all()
        if data_inicio:
            agendamentos = agendamentos.filter(data_agendamento__gte=data_inicio)
        if data_fim:
            agendamentos = agendamentos.filter(data_agendamento__lte=data_fim)
        for agendamento in agendamentos:
            atividades.append(
                {
                    "data": agendamento.data_agendamento,
                    "data_ordem": normalize_datetime(
                        datetime.combine(agendamento.data_agendamento, datetime.min.time())
                    ),
                    "atividade": "Novo Agendamento",
                    "tecnico": "Sistema",
                    "egresso": agendamento.nome_atendido,
                    "acoes": {
                        "ver": f"/agendamentos/{agendamento.id}/ver",
                        "editar": f"/agendamentos/{agendamento.id}/editar",
                        "excluir": f"/agendamentos/{agendamento.id}/excluir",
                    },
                }
            )

    if should_include("atendimentos"):
        atendimentos = Atendimento.objects.all()
        if data_inicio:
            atendimentos = atendimentos.filter(data_atendimento__gte=data_inicio)
        if data_fim:
            atendimentos = atendimentos.filter(data_atendimento__lte=data_fim)
        for atendimento in atendimentos:
            atividades.append(
                {
                    "data": atendimento.data_atendimento,
                    "data_ordem": normalize_datetime(
                        datetime.combine(atendimento.data_atendimento, datetime.min.time())
                    ),
                    "atividade": "Novo Atendimento",
                    "tecnico": atendimento.profissional_responsavel or "Sistema",
                    "egresso": atendimento.nome_pessoa_atendida,
                    "acoes": {
                        "ver": f"/atendimentos/{atendimento.id}/ver",
                        "editar": f"/atendimentos/{atendimento.id}/editar",
                        "excluir": f"/atendimentos/{atendimento.id}/excluir",
                    },
                }
            )

    if should_include("anotacoes"):
        lembretes = Lembrete.objects.all()
        if data_inicio:
            lembretes = lembretes.filter(data_criacao__date__gte=data_inicio)
        if data_fim:
            lembretes = lembretes.filter(data_criacao__date__lte=data_fim)
        for lembrete in lembretes:
            atividades.append(
                {
                    "data": lembrete.data_criacao,
                    "data_ordem": normalize_datetime(lembrete.data_criacao),
                    "atividade": "Anotação",
                    "tecnico": getattr(lembrete.criado_por, "username", "Sistema"),
                    "egresso": lembrete.cadastro.nome,
                    "acoes": {
                        "ver": f"/cadastro/perfil/{lembrete.cadastro.id}",
                        "editar": "/anotacoes/editar/?cadastro_id={}".format(lembrete.cadastro.id),
                        "excluir": None,
                    },
                }
            )

    atividades.sort(key=lambda item: item["data_ordem"], reverse=True)

    context = {
        "atividades": atividades,
        "total_atividades": len(atividades),
        "filtro_tipo": tipo,
        "data_inicio": data_inicio_raw,
        "data_fim": data_fim_raw,
    }
    return render(request, "core/minhas_atividades.html", context)


def logout_view(request):
    """Log the user out and redirect to login."""
    logout(request)
    return redirect("login")


def is_admin_user(user) -> bool:
    """Check if the user can access admin pages."""
    return user.is_authenticated and user.is_staff


@login_required
@user_passes_test(is_admin_user)
def admin_users_list_view(request):
    """List system users for admin management."""
    User = get_user_model()
    users = User.objects.all().order_by("username")
    profiles = {profile.user_id: profile for profile in UserProfile.objects.filter(user__in=users)}
    return render(
        request,
        "core/admin_users_list.html",
        {
            "users": users,
            "profiles": profiles,
        },
    )


@login_required
@user_passes_test(is_admin_user)
def admin_user_add_view(request):
    """Create a new system user."""
    if request.method == "POST":
        form = AdminUserCreateForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Usuário criado com sucesso.")
                return redirect("admin-users-list")
            except IntegrityError:
                form.add_error("username", "Nome de usuário já existe.")
    else:
        form = AdminUserCreateForm()

    return render(request, "core/admin_user_add.html", {"form": form})


@login_required
@user_passes_test(is_admin_user)
def admin_user_detail_view(request, user_id: int):
    """Show details of a specific user."""
    User = get_user_model()
    user = get_object_or_404(User, id=user_id)
    profile = UserProfile.objects.filter(user=user).first()
    return render(request, "core/admin_user_detail.html", {"user_obj": user, "profile": profile})


@login_required
@user_passes_test(is_admin_user)
def admin_user_toggle_active_view(request, user_id: int):
    """Activate or deactivate a user account."""
    User = get_user_model()
    user = get_object_or_404(User, id=user_id)
    if user.id != request.user.id:
        user.is_active = not user.is_active
        user.save()
        messages.success(
            request,
            "Usuário ativado com sucesso." if user.is_active else "Usuário desativado com sucesso.",
        )
    return redirect("admin-users-list")


@login_required
@user_passes_test(is_admin_user)
def admin_user_delete_view(request, user_id: int):
    """Delete a user account after confirmation."""
    User = get_user_model()
    user = get_object_or_404(User, id=user_id)
    if request.method == "POST" and user.id != request.user.id:
        user.delete()
        messages.success(request, "Usuário excluído com sucesso.")
        return redirect("admin-users-list")
    return render(request, "core/admin_user_delete.html", {"user_obj": user})


@login_required
def api_buscar_cadastros_view(request):
    """API endpoint to search cadastros by name (for autocomplete)."""
    query = request.GET.get("q", "").strip()
    
    if len(query) < 2:
        return JsonResponse({"results": []})
    
    cadastros = Cadastro.objects.filter(
        nome__icontains=query,
        status="ativo"
    ).values("id", "nome")[:10]
    
    results = [
        {"id": c["id"], "text": c["nome"]}
        for c in cadastros
    ]
    
    return JsonResponse({"results": results})


@login_required
def agendamentos_dashboard_view(request):
    """Render the agendamentos dashboard with search and lists."""
    query = request.GET.get("q", "").strip()
    agendamentos = Agendamento.objects.all().order_by("-data_agendamento")
    
    if query:
        agendamentos = agendamentos.filter(nome_atendido__icontains=query)
    
    # Add cadastro object to each agendamento for linking
    agendamentos_with_cadastro = []
    for agendamento in agendamentos:
        cadastro = Cadastro.objects.filter(nome=agendamento.nome_atendido).first()
        agendamentos_with_cadastro.append({
            'agendamento': agendamento,
            'cadastro': cadastro
        })
    
    context = {
        "query": query,
        "agendamentos_with_cadastro": agendamentos_with_cadastro,
        "total": agendamentos.count(),
    }
    return render(request, "core/agendamentos_dashboard.html", context)


def login_view(request):
    """Render login page and authenticate user credentials."""
    # Read redirect target from query string first (GET).
    redirect_to = request.GET.get("redirect", "/home/")
    error = None

    if request.method == "POST":
        # For POST, respect hidden redirect field from the form.
        redirect_to = request.POST.get("redirect", "/home/")
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")

        # Delegate authentication to Django's auth system.
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect(redirect_to)

        # Dev helper: ensure default admin exists if credentials match env.
        if settings.DEBUG:
            from .default_admin import ensure_default_admin_exists, get_default_admin_data

            data = get_default_admin_data()
            if data and username == data["username"] and password == data["password"]:
                ensure_default_admin_exists()
                user = authenticate(request, username=username, password=password)
                if user is not None:
                    login(request, user)
                    return redirect(redirect_to)
        error = "Usuário ou senha inválidos."

    context = {
        "redirect_to": redirect_to,
        "error": error,
    }
    return render(request, "core/login.html", context)


# ────────────────────────────────────────────────────────
#  Relatórios
# ────────────────────────────────────────────────────────


@login_required
def relatorios_view(request):
    """Render the reports page with filters."""
    context = {
        "total_cadastros": Cadastro.objects.count(),
        "total_familiares": Familiar.objects.count(),
        "total_atendimentos": Atendimento.objects.count(),
        "total_agendamentos": Agendamento.objects.count(),
    }
    return render(request, "core/relatorios.html", context)


def _apply_cadastro_filters(qs, params):
    """Apply query-string filters to a Cadastro queryset."""
    if params.get("status_cadastro"):
        qs = qs.filter(status=params["status_cadastro"])
    if params.get("data_inicio"):
        qs = qs.filter(data_cadastro__gte=params["data_inicio"])
    if params.get("data_fim"):
        qs = qs.filter(data_cadastro__lte=params["data_fim"])
    if params.get("sexo_biologico"):
        qs = qs.filter(sexo_biologico=params["sexo_biologico"])
    if params.get("etnia"):
        qs = qs.filter(identidade_etnico_racial=params["etnia"])
    if params.get("grau_instrucao"):
        qs = qs.filter(grau_instrucao=params["grau_instrucao"])
    if params.get("status_ocupacional"):
        qs = qs.filter(status_ocupacional=params["status_ocupacional"])
    if params.get("identidade_genero"):
        qs = qs.filter(identidade_genero=params["identidade_genero"])
    if params.get("experiencia_trabalho"):
        qs = qs.filter(experiencia_trabalho=params["experiencia_trabalho"])
    if params.get("tipo_ocupacao"):
        qs = qs.filter(tipo_ocupacao=params["tipo_ocupacao"])
    if params.get("deficiencia"):
        qs = qs.filter(deficiencias__icontains=params["deficiencia"])
    if params.get("possui_comorbidade"):
        qs = qs.filter(possui_comorbidade=params["possui_comorbidade"])
    if params.get("uso_substancias_psicoativas"):
        qs = qs.filter(
            uso_substancias_psicoativas=params["uso_substancias_psicoativas"]
        )
    if params.get("zona_cidade"):
        qs = qs.filter(zona_cidade=params["zona_cidade"])
    if params.get("nome"):
        qs = qs.filter(nome__icontains=params["nome"])
    return qs


def _report_querysets(params):
    """Return consistently filtered querysets used by every report section."""
    cadastros = _apply_cadastro_filters(
        Cadastro.objects.all().order_by("nome"), params
    )
    familiares = Familiar.objects.all().order_by("nome")
    atendimentos = Atendimento.objects.all().order_by("-data_atendimento")
    agendamentos = Agendamento.objects.all().order_by("-data_agendamento")

    # Each report section uses its own registration date. A Familiar may be
    # avulso (without Cadastro FK), so filtering it through Cadastro would
    # incorrectly remove most of the records shown on the Familiares page.
    if params.get("data_inicio"):
        familiares = familiares.filter(data_criacao__date__gte=params["data_inicio"])
    if params.get("data_fim"):
        familiares = familiares.filter(data_criacao__date__lte=params["data_fim"])
    if params.get("sexo_biologico"):
        familiares = familiares.filter(sexo_biologico=params["sexo_biologico"])
    if params.get("etnia"):
        familiares = familiares.filter(identidade_etnico_racial=params["etnia"])
    for key in (
        "identidade_genero",
        "experiencia_trabalho",
        "tipo_ocupacao",
        "grau_instrucao",
        "possui_comorbidade",
        "uso_substancias_psicoativas",
    ):
        if params.get(key):
            familiares = familiares.filter(**{key: params[key]})
    if params.get("deficiencia"):
        familiares = familiares.filter(deficiencias__icontains=params["deficiencia"])
    if params.get("nome"):
        nome = params["nome"]
        familiares = familiares.filter(
            db_models.Q(nome__icontains=nome)
            | db_models.Q(nome_interno_referencia__icontains=nome)
            | db_models.Q(cadastro__nome__icontains=nome)
        )

    # These attributes only exist on a linked egresso. When selected, avulsos
    # cannot be evaluated and only relatives of matching cadastros are valid.
    linked_filter_keys = (
        "status_cadastro", "status_ocupacional", "zona_cidade",
    )
    if any(params.get(key) for key in linked_filter_keys):
        familiares = familiares.filter(cadastro__in=cadastros)

    # Activities do not have a foreign key to Cadastro, but their own date and
    # attended-person name can still honor the equivalent report filters.
    if params.get("data_inicio"):
        atendimentos = atendimentos.filter(data_atendimento__gte=params["data_inicio"])
        agendamentos = agendamentos.filter(data_agendamento__gte=params["data_inicio"])
    if params.get("data_fim"):
        atendimentos = atendimentos.filter(data_atendimento__lte=params["data_fim"])
        agendamentos = agendamentos.filter(data_agendamento__lte=params["data_fim"])
    if params.get("nome"):
        atendimentos = atendimentos.filter(nome_pessoa_atendida__icontains=params["nome"])
        agendamentos = agendamentos.filter(nome_atendido__icontains=params["nome"])

    return cadastros, familiares, atendimentos, agendamentos


AGE_BANDS = (
    ("Menor de 18", None, 17),
    ("18 a 24", 18, 24),
    ("25 a 29", 25, 29),
    ("30 a 34", 30, 34),
    ("35 a 59", 35, 59),
    ("60 a 74", 60, 74),
    ("75 ou mais", 75, None),
)


def _age_on(birth_date, reference_date=None):
    if not birth_date:
        return None
    reference_date = reference_date or timezone.localdate()
    return (
        reference_date.year
        - birth_date.year
        - ((reference_date.month, reference_date.day) < (birth_date.month, birth_date.day))
    )


def _age_distribution(objects):
    counts = {label: 0 for label, _minimum, _maximum in AGE_BANDS}
    missing = 0
    for obj in objects:
        age = _age_on(obj.data_nascimento)
        if age is None:
            missing += 1
            continue
        for label, minimum, maximum in AGE_BANDS:
            if (minimum is None or age >= minimum) and (maximum is None or age <= maximum):
                counts[label] += 1
                break
    rows = list(counts.items())
    if missing:
        rows.append(("Data de nascimento não informada", missing))
    return rows


def _age_band_label(birth_date):
    age = _age_on(birth_date)
    if age is None:
        return "Não informada"
    for label, minimum, maximum in AGE_BANDS:
        if (minimum is None or age >= minimum) and (maximum is None or age <= maximum):
            return label
    return "Não informada"


def _choice_distribution(queryset, field, choices):
    rows = [(label, queryset.filter(**{field: value}).count()) for value, label in choices]
    blank_count = queryset.filter(**{field: ""}).count()
    if blank_count:
        rows.append(("Não informado", blank_count))
    return rows


def _multiple_choice_distribution(objects, field, choices):
    counts = {value: 0 for value, _label in choices}
    unreported = 0
    for obj in objects:
        selected = {
            item.strip() for item in (getattr(obj, field, "") or "").split(",") if item.strip()
        }
        if not selected:
            unreported += 1
        for value in selected:
            if value in counts:
                counts[value] += 1
    rows = [(label, counts[value]) for value, label in choices]
    if unreported:
        rows.append(("Sem deficiência informada", unreported))
    return rows


def _text_distribution(objects, field, empty_label="Não informado"):
    counts = {}
    for obj in objects:
        value = (getattr(obj, field, "") or "").strip() or empty_label
        counts[value] = counts.get(value, 0) + 1
    return sorted(counts.items(), key=lambda item: item[0].casefold())


def _deficiencias_display(obj):
    labels = dict(Cadastro.DEFICIENCIA_CHOICES)
    selected = [
        labels.get(item.strip(), item.strip())
        for item in (obj.deficiencias or "").split(",")
        if item.strip()
    ]
    return ", ".join(selected)


@login_required
def relatorios_pdf_view(request):
    """Generate a PDF report based on selected sections and filters."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import io
    from xml.sax.saxutils import escape

    params = request.GET
    secoes = params.getlist("secao")
    if not secoes:
        secoes = ["cadastros", "familiares", "atendimentos", "agendamentos", "quantitativos"]

    cadastros, familiares_qs, atendimentos_qs, agendamentos_qs = _report_querysets(params)

    # ── Build PDF ──
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=20 * mm, rightMargin=20 * mm,
        topMargin=20 * mm, bottomMargin=20 * mm,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        "SectionTitle", parent=styles["Heading2"],
        textColor=colors.HexColor("#0066FF"), spaceBefore=16, spaceAfter=8,
    ))
    styles.add(ParagraphStyle(
        "CellText", parent=styles["Normal"], fontSize=8, leading=10,
    ))
    styles.add(ParagraphStyle(
        "SmallBold", parent=styles["Normal"], fontSize=8, leading=10,
        textColor=colors.white,
    ))

    story = []

    # ── Header ──
    header_style = ParagraphStyle(
        "ReportHeader", parent=styles["Title"],
        textColor=colors.HexColor("#0066FF"), alignment=TA_CENTER,
    )
    story.append(Paragraph("ESPI — Relatório do Sistema", header_style))
    story.append(Spacer(1, 4 * mm))

    subtitle = ParagraphStyle(
        "Subtitle", parent=styles["Normal"],
        alignment=TA_CENTER, textColor=colors.grey, fontSize=10,
    )
    now = timezone.localtime(timezone.now())
    story.append(Paragraph(
        f"Gerado em {now.strftime('%d/%m/%Y')} às {now.strftime('%H:%M')} por {request.user.username}",
        subtitle,
    ))
    story.append(Spacer(1, 8 * mm))

    # Helper for tables
    base_table_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0066FF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ])

    cell = styles["CellText"]

    def make_para(text):
        return Paragraph(escape(str(text)) if text else "—", cell)

    def make_lines(*lines):
        safe_lines = [escape(str(line)) for line in lines if line]
        return Paragraph("<br/>".join(safe_lines) if safe_lines else "—", cell)

    def header_para(text):
        return Paragraph(str(text), styles["SmallBold"])

    # ── Quantitativos ──
    if "quantitativos" in secoes:
        quant_title = Paragraph("Resumo Quantitativo", styles["SectionTitle"])
        total_cad = cadastros.count()
        ativos = cadastros.filter(status="ativo").count()
        arquivados = cadastros.filter(status="arquivado").count()
        total_fam = familiares_qs.count()
        fam_vinculados = familiares_qs.filter(cadastro__isnull=False).count()
        fam_avulsos = familiares_qs.filter(cadastro__isnull=True).count()
        total_atd = atendimentos_qs.count()
        total_agd = agendamentos_qs.count()

        summary_data = [
            [header_para("Indicador"), header_para("Quantidade")],
            [make_para("Total de Cadastros / Egressos"), make_para(str(total_cad))],
            [make_para("  ↳ Ativos"), make_para(str(ativos))],
            [make_para("  ↳ Arquivados"), make_para(str(arquivados))],
            [make_para("Total de Familiares"), make_para(str(total_fam))],
            [make_para("  ↳ Vinculados a cadastro"), make_para(str(fam_vinculados))],
            [make_para("  ↳ Avulsos (sem vínculo)"), make_para(str(fam_avulsos))],
            [make_para("Total de Atendimentos"), make_para(str(total_atd))],
            [make_para("Total de Agendamentos"), make_para(str(total_agd))],
        ]
        tbl = Table(summary_data, colWidths=[120 * mm, 40 * mm])
        tbl.setStyle(base_table_style)
        story.append(KeepTogether([quant_title, tbl]))
        story.append(Spacer(1, 6 * mm))

        def add_distribution(title, category_label, rows):
            title_para = Paragraph(title, styles["SectionTitle"])
            dist_data = [[header_para(category_label), header_para("Quantidade")]]
            for display, count in rows:
                dist_data.append([make_para(display), make_para(str(count))])
            tbl = Table(dist_data, colWidths=[120 * mm, 40 * mm])
            tbl.setStyle(base_table_style)
            story.append(KeepTogether([title_para, tbl, Spacer(1, 4 * mm)]))

        egresso_distributions = [
            ("Pessoas Egressas por Faixa Etária", "Faixa etária", _age_distribution(cadastros)),
            ("Pessoas Egressas por Gênero", "Gênero", _choice_distribution(
                cadastros, "identidade_genero", Cadastro.IDENTIDADE_GENERO_CHOICES
            )),
            ("Pessoas Egressas por Deficiência", "Tipo de deficiência", _multiple_choice_distribution(
                cadastros, "deficiencias", Cadastro.DEFICIENCIA_CHOICES
            )),
            ("Pessoas Egressas por Escolarização", "Grau de instrução", _choice_distribution(
                cadastros, "grau_instrucao", Cadastro.GRAU_INSTRUCAO_CHOICES
            )),
            ("Pessoas Egressas por Série Final de Estudo", "Série final", _text_distribution(
                cadastros, "serie_concluida"
            )),
            ("Pessoas Egressas por Tipo de Ocupação", "Tipo de ocupação", _choice_distribution(
                cadastros, "tipo_ocupacao", Cadastro.TIPO_OCUPACAO_CHOICES
            )),
            ("Pessoas Egressas por Experiência de Trabalho", "Experiência", _choice_distribution(
                cadastros, "experiencia_trabalho", Cadastro.EXPERIENCIA_TRABALHO_CHOICES
            )),
            ("Pessoas Egressas por Comorbidade", "Possui comorbidade", _choice_distribution(
                cadastros, "possui_comorbidade", Cadastro.COMORBIDADE_CHOICES
            )),
            ("Pessoas Egressas por Uso de Substâncias Psicoativas", "Uso declarado", _choice_distribution(
                cadastros, "uso_substancias_psicoativas", Cadastro.USO_SUBSTANCIAS_CHOICES
            )),
        ]
        familiar_distributions = [
            ("Familiares por Faixa Etária", "Faixa etária", _age_distribution(familiares_qs)),
            ("Familiares por Gênero", "Gênero", _choice_distribution(
                familiares_qs, "identidade_genero", Cadastro.IDENTIDADE_GENERO_CHOICES
            )),
            ("Familiares por Deficiência", "Tipo de deficiência", _multiple_choice_distribution(
                familiares_qs, "deficiencias", Cadastro.DEFICIENCIA_CHOICES
            )),
            ("Familiares por Escolarização", "Grau de instrução", _choice_distribution(
                familiares_qs, "grau_instrucao", Cadastro.GRAU_INSTRUCAO_CHOICES
            )),
            ("Familiares por Série Final de Estudo", "Série final", _text_distribution(
                familiares_qs, "serie_concluida"
            )),
            ("Familiares por Tipo de Ocupação", "Tipo de ocupação", _choice_distribution(
                familiares_qs, "tipo_ocupacao", Cadastro.TIPO_OCUPACAO_CHOICES
            )),
            ("Familiares por Experiência de Trabalho", "Experiência", _choice_distribution(
                familiares_qs, "experiencia_trabalho", Cadastro.EXPERIENCIA_TRABALHO_CHOICES
            )),
            ("Familiares por Comorbidade", "Possui comorbidade", _choice_distribution(
                familiares_qs, "possui_comorbidade", Cadastro.COMORBIDADE_CHOICES
            )),
            ("Familiares por Uso de Substâncias Psicoativas", "Uso declarado", _choice_distribution(
                familiares_qs, "uso_substancias_psicoativas", Cadastro.USO_SUBSTANCIAS_CHOICES
            )),
        ]
        for title, category_label, rows in egresso_distributions + familiar_distributions:
            add_distribution(title, category_label, rows)

    # ── Cadastros ──
    if "cadastros" in secoes:
        cad_title = Paragraph(f"Cadastros / Egressos ({cadastros.count()})", styles["SectionTitle"])
        cad_data = [[
            header_para("Nome"), header_para("CPF"), header_para("Status"),
            header_para("Faixa etária / Gênero"), header_para("Trabalho / Ocupação"),
            header_para("Escolaridade / Série"), header_para("Saúde e substâncias"),
        ]]
        for c in cadastros:
            cad_data.append([
                make_para(c.nome),
                make_para(c.cpf_numero),
                make_para(c.get_status_display()),
                make_lines(
                    f"Faixa: {_age_band_label(c.data_nascimento)}",
                    f"Gênero: {c.get_identidade_genero_display()}" if c.identidade_genero else "",
                ),
                make_lines(
                    f"Experiência: {c.get_experiencia_trabalho_display()}" if c.experiencia_trabalho else "",
                    f"Ocupação: {c.get_tipo_ocupacao_display()}" if c.tipo_ocupacao else "",
                ),
                make_lines(
                    c.get_grau_instrucao_display() if c.grau_instrucao else "",
                    f"Série final: {c.serie_concluida}" if c.serie_concluida else "",
                ),
                make_lines(
                    f"Deficiência: {_deficiencias_display(c)}" if c.deficiencias else "",
                    f"Comorbidade: {c.get_possui_comorbidade_display()}" if c.possui_comorbidade else "",
                    c.comorbidades,
                    f"Substâncias: {c.get_uso_substancias_psicoativas_display()}" if c.uso_substancias_psicoativas else "",
                    c.substancias_psicoativas,
                ),
            ])
        if len(cad_data) == 1:
            cad_data.append([make_para("Nenhum registro encontrado")] + [make_para("")] * 6)
        tbl = Table(cad_data, colWidths=[28 * mm, 20 * mm, 16 * mm, 25 * mm, 27 * mm, 27 * mm, 27 * mm])
        tbl.setStyle(base_table_style)
        story.append(Spacer(1, 6 * mm))
        story.append(KeepTogether([cad_title, tbl]))

    # ── Familiares ──
    if "familiares" in secoes:
        fam_title = Paragraph(f"Familiares ({familiares_qs.count()})", styles["SectionTitle"])
        fam_data = [[
            header_para("Nome"), header_para("Parentesco"),
            header_para("Faixa / Gênero"), header_para("Trabalho / Ocupação"),
            header_para("Escolaridade / Série"), header_para("Saúde e substâncias"),
        ]]
        for f in familiares_qs:
            fam_data.append([
                make_para(f.nome),
                make_para(f.parentesco),
                make_lines(
                    f"Faixa: {_age_band_label(f.data_nascimento)}",
                    f"Gênero: {f.get_identidade_genero_display()}" if f.identidade_genero else "",
                ),
                make_lines(
                    f"Experiência: {f.get_experiencia_trabalho_display()}" if f.experiencia_trabalho else "",
                    f"Tipo: {f.get_tipo_ocupacao_display()}" if f.tipo_ocupacao else "",
                    f"Ocupação: {f.ocupacao}" if f.ocupacao else "",
                ),
                make_lines(
                    f.get_grau_instrucao_display() if f.grau_instrucao else "",
                    f"Série final: {f.serie_concluida}" if f.serie_concluida else "",
                ),
                make_lines(
                    f"Deficiência: {_deficiencias_display(f)}" if f.deficiencias else "",
                    f"Comorbidade: {f.get_possui_comorbidade_display()}" if f.possui_comorbidade else "",
                    f.comorbidades,
                    f"Substâncias: {f.get_uso_substancias_psicoativas_display()}" if f.uso_substancias_psicoativas else "",
                    f.substancias_psicoativas,
                ),
            ])
        if len(fam_data) == 1:
            fam_data.append([make_para("Nenhum registro encontrado")] + [make_para("")] * 5)
        tbl = Table(fam_data, colWidths=[28 * mm, 20 * mm, 27 * mm, 30 * mm, 28 * mm, 27 * mm])
        tbl.setStyle(base_table_style)
        story.append(Spacer(1, 6 * mm))
        story.append(KeepTogether([fam_title, tbl]))

    # ── Atendimentos ──
    if "atendimentos" in secoes:
        atd_title = Paragraph(f"Atendimentos ({atendimentos_qs.count()})", styles["SectionTitle"])
        atd_data = [[
            header_para("Data"), header_para("Pessoa Atendida"),
            header_para("Tipo"), header_para("Local"),
            header_para("Motivo"), header_para("Status"),
        ]]
        for a in atendimentos_qs:
            atd_data.append([
                make_para(a.data_atendimento.strftime("%d/%m/%Y") if a.data_atendimento else ""),
                make_para(a.nome_pessoa_atendida),
                make_para(a.get_tipo_atendimento_display() if a.tipo_atendimento else ""),
                make_para(a.get_local_atendimento_display() if a.local_atendimento else ""),
                make_para(a.get_motivo_procura_display() if a.motivo_procura else ""),
                make_para(a.get_status_display() if a.status else ""),
            ])
        if len(atd_data) == 1:
            atd_data.append([make_para("Nenhum registro encontrado")] + [make_para("")] * 5)
        tbl = Table(atd_data, colWidths=[22 * mm, 35 * mm, 25 * mm, 28 * mm, 25 * mm, 20 * mm])
        tbl.setStyle(base_table_style)
        story.append(Spacer(1, 6 * mm))
        story.append(KeepTogether([atd_title, tbl]))

    # ── Agendamentos ──
    if "agendamentos" in secoes:
        agd_title = Paragraph(f"Agendamentos ({agendamentos_qs.count()})", styles["SectionTitle"])
        agd_data = [[
            header_para("Data"), header_para("Horário"),
            header_para("Nome"), header_para("Tipo"),
            header_para("Observações"),
        ]]
        for ag in agendamentos_qs:
            agd_data.append([
                make_para(ag.data_agendamento.strftime("%d/%m/%Y") if ag.data_agendamento else ""),
                make_para(ag.horario_atendimento),
                make_para(ag.nome_atendido),
                make_para(ag.get_tipo_agendamento_display() if ag.tipo_agendamento else ""),
                make_para(ag.observacoes[:80] if ag.observacoes else ""),
            ])
        if len(agd_data) == 1:
            agd_data.append([make_para("Nenhum registro encontrado")] + [make_para("")] * 4)
        tbl = Table(agd_data, colWidths=[24 * mm, 18 * mm, 40 * mm, 28 * mm, 50 * mm])
        tbl.setStyle(base_table_style)
        story.append(Spacer(1, 6 * mm))
        story.append(KeepTogether([agd_title, tbl]))

    # Build
    doc.build(story)
    buf.seek(0)

    response = HttpResponse(buf.read(), content_type="application/pdf")
    filename = f"relatorio_espi_{now.strftime('%Y%m%d_%H%M')}.pdf"
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


@login_required
def relatorios_excel_view(request):
    """Generate a formatted XLSX workbook using the same filters as the PDF."""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    params = request.GET
    secoes = params.getlist("secao")
    if not secoes:
        secoes = ["cadastros", "familiares", "atendimentos", "agendamentos", "quantitativos"]

    cadastros, familiares_qs, atendimentos_qs, agendamentos_qs = _report_querysets(params)
    cadastros = list(cadastros)
    familiares = list(familiares_qs)
    atendimentos = list(atendimentos_qs)
    agendamentos = list(agendamentos_qs)

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = "Relatório do Sistema ESPI"
    workbook.properties.subject = "Dados de pessoas egressas, familiares e atividades"
    workbook.properties.creator = request.user.get_username()
    workbook.properties.description = "Exportação gerada pelo ESPI com os filtros selecionados."

    blue = "0B5ED7"
    dark_blue = "073B88"
    light_blue = "DCEBFF"
    lighter_blue = "F3F7FC"
    border_color = "B8CCE4"
    white = "FFFFFF"
    muted = "5D6B7A"
    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    choice_maps = {
        "status_cadastro": dict(Cadastro.STATUS_CHOICES),
        "sexo_biologico": dict(Cadastro.SEXO_BIOLOGICO_CHOICES),
        "identidade_genero": dict(Cadastro.IDENTIDADE_GENERO_CHOICES),
        "etnia": dict(Cadastro.ETNIA_CHOICES),
        "grau_instrucao": dict(Cadastro.GRAU_INSTRUCAO_CHOICES),
        "status_ocupacional": dict(Cadastro.STATUS_OCUPACIONAL_CHOICES),
        "experiencia_trabalho": dict(Cadastro.EXPERIENCIA_TRABALHO_CHOICES),
        "tipo_ocupacao": dict(Cadastro.TIPO_OCUPACAO_CHOICES),
        "deficiencia": dict(Cadastro.DEFICIENCIA_CHOICES),
        "possui_comorbidade": dict(Cadastro.COMORBIDADE_CHOICES),
        "uso_substancias_psicoativas": dict(Cadastro.USO_SUBSTANCIAS_CHOICES),
        "zona_cidade": dict(Cadastro.ZONA_CIDADE_CHOICES),
    }
    filter_names = {
        "status_cadastro": "Status",
        "data_inicio": "Data inicial",
        "data_fim": "Data final",
        "sexo_biologico": "Sexo biológico",
        "identidade_genero": "Gênero",
        "etnia": "Etnia",
        "grau_instrucao": "Escolaridade",
        "experiencia_trabalho": "Experiência de trabalho",
        "tipo_ocupacao": "Tipo de ocupação",
        "deficiencia": "Deficiência",
        "possui_comorbidade": "Comorbidade",
        "uso_substancias_psicoativas": "Uso de substâncias",
        "status_ocupacional": "Status ocupacional",
        "zona_cidade": "Zona",
        "nome": "Nome",
    }
    active_filters = []
    for key, label in filter_names.items():
        value = params.get(key)
        if value:
            display = choice_maps.get(key, {}).get(value, value)
            active_filters.append(f"{label}: {display}")
    filters_text = "Filtros: " + (" | ".join(active_filters) if active_filters else "nenhum")

    def safe_value(value):
        if value is None:
            return ""
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
            return "'" + value
        return value

    def add_sheet(name, title, headers, rows, tab_color=blue):
        sheet = workbook.create_sheet(name)
        sheet.sheet_properties.tabColor = tab_color
        sheet.sheet_view.showGridLines = False
        last_column = get_column_letter(len(headers))

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = sheet.cell(1, 1, title)
        title_cell.fill = PatternFill("solid", fgColor=dark_blue)
        title_cell.font = Font(color=white, bold=True, size=16)
        title_cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[1].height = 30

        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        info_cell = sheet.cell(
            2,
            1,
            f"Gerado em {timezone.localtime(timezone.now()).strftime('%d/%m/%Y às %H:%M')} por "
            f"{request.user.get_username()} — {filters_text}",
        )
        info_cell.font = Font(color=muted, italic=True, size=9)
        info_cell.alignment = Alignment(wrap_text=True, vertical="center")
        sheet.row_dimensions[2].height = 30

        for column, header in enumerate(headers, 1):
            cell = sheet.cell(3, column, header)
            cell.fill = PatternFill("solid", fgColor=blue)
            cell.font = Font(color=white, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        sheet.row_dimensions[3].height = 32

        for row_index, row in enumerate(rows, 4):
            for column, value in enumerate(row, 1):
                cell = sheet.cell(row_index, column, safe_value(value))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border
                if row_index % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=lighter_blue)
                if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
                    cell.number_format = "dd/mm/yyyy"
            sheet.row_dimensions[row_index].height = 28

        sheet.freeze_panes = "A4"
        sheet.auto_filter.ref = f"A3:{last_column}{max(sheet.max_row, 3)}"
        if rows:
            table = Table(
                displayName=f"TabelaESPI{len(workbook.worksheets)}",
                ref=f"A3:{last_column}{sheet.max_row}",
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)
        sheet.print_title_rows = "1:3"
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        for column in range(1, len(headers) + 1):
            values = [str(sheet.cell(row, column).value or "") for row in range(3, sheet.max_row + 1)]
            width = min(max(max((len(value) for value in values), default=10) + 2, 12), 42)
            sheet.column_dimensions[get_column_letter(column)].width = width
        return sheet

    if "quantitativos" in secoes:
        summary_rows = [
            ["Totais", "Cadastros / Egressos", "Total", len(cadastros)],
            ["Totais", "Cadastros / Egressos", "Ativos", sum(c.status == "ativo" for c in cadastros)],
            ["Totais", "Cadastros / Egressos", "Arquivados", sum(c.status == "arquivado" for c in cadastros)],
            ["Totais", "Familiares", "Total", len(familiares)],
            ["Totais", "Familiares", "Vinculados", sum(bool(f.cadastro_id) for f in familiares)],
            ["Totais", "Familiares", "Avulsos", sum(not f.cadastro_id for f in familiares)],
            ["Totais", "Atendimentos", "Total", len(atendimentos)],
            ["Totais", "Agendamentos", "Total", len(agendamentos)],
        ]
        distributions = [
            ("Egressos", "Faixa etária", _age_distribution(cadastros)),
            ("Egressos", "Gênero", _choice_distribution(
                Cadastro.objects.filter(pk__in=[c.pk for c in cadastros]),
                "identidade_genero",
                Cadastro.IDENTIDADE_GENERO_CHOICES,
            )),
            ("Egressos", "Deficiência", _multiple_choice_distribution(
                cadastros, "deficiencias", Cadastro.DEFICIENCIA_CHOICES
            )),
            ("Egressos", "Escolaridade", _choice_distribution(
                Cadastro.objects.filter(pk__in=[c.pk for c in cadastros]),
                "grau_instrucao",
                Cadastro.GRAU_INSTRUCAO_CHOICES,
            )),
            ("Egressos", "Série final", _text_distribution(cadastros, "serie_concluida")),
            ("Egressos", "Tipo de ocupação", _choice_distribution(
                Cadastro.objects.filter(pk__in=[c.pk for c in cadastros]),
                "tipo_ocupacao",
                Cadastro.TIPO_OCUPACAO_CHOICES,
            )),
            ("Egressos", "Experiência de trabalho", _choice_distribution(
                Cadastro.objects.filter(pk__in=[c.pk for c in cadastros]),
                "experiencia_trabalho",
                Cadastro.EXPERIENCIA_TRABALHO_CHOICES,
            )),
            ("Egressos", "Comorbidade", _choice_distribution(
                Cadastro.objects.filter(pk__in=[c.pk for c in cadastros]),
                "possui_comorbidade",
                Cadastro.COMORBIDADE_CHOICES,
            )),
            ("Egressos", "Uso de substâncias", _choice_distribution(
                Cadastro.objects.filter(pk__in=[c.pk for c in cadastros]),
                "uso_substancias_psicoativas",
                Cadastro.USO_SUBSTANCIAS_CHOICES,
            )),
            ("Familiares", "Faixa etária", _age_distribution(familiares)),
            ("Familiares", "Gênero", _choice_distribution(
                Familiar.objects.filter(pk__in=[f.pk for f in familiares]),
                "identidade_genero",
                Cadastro.IDENTIDADE_GENERO_CHOICES,
            )),
            ("Familiares", "Deficiência", _multiple_choice_distribution(
                familiares, "deficiencias", Cadastro.DEFICIENCIA_CHOICES
            )),
            ("Familiares", "Escolaridade", _choice_distribution(
                Familiar.objects.filter(pk__in=[f.pk for f in familiares]),
                "grau_instrucao",
                Cadastro.GRAU_INSTRUCAO_CHOICES,
            )),
            ("Familiares", "Série final", _text_distribution(familiares, "serie_concluida")),
            ("Familiares", "Tipo de ocupação", _choice_distribution(
                Familiar.objects.filter(pk__in=[f.pk for f in familiares]),
                "tipo_ocupacao",
                Cadastro.TIPO_OCUPACAO_CHOICES,
            )),
            ("Familiares", "Experiência de trabalho", _choice_distribution(
                Familiar.objects.filter(pk__in=[f.pk for f in familiares]),
                "experiencia_trabalho",
                Cadastro.EXPERIENCIA_TRABALHO_CHOICES,
            )),
            ("Familiares", "Comorbidade", _choice_distribution(
                Familiar.objects.filter(pk__in=[f.pk for f in familiares]),
                "possui_comorbidade",
                Cadastro.COMORBIDADE_CHOICES,
            )),
            ("Familiares", "Uso de substâncias", _choice_distribution(
                Familiar.objects.filter(pk__in=[f.pk for f in familiares]),
                "uso_substancias_psicoativas",
                Cadastro.USO_SUBSTANCIAS_CHOICES,
            )),
        ]
        for group, indicator, rows in distributions:
            summary_rows.extend([group, indicator, category, quantity] for category, quantity in rows)
        summary_sheet = add_sheet(
            "Resumo",
            "ESPI — Resumo quantitativo",
            ["Grupo", "Indicador", "Categoria", "Quantidade"],
            summary_rows,
            "70AD47",
        )
        summary_sheet.column_dimensions["A"].width = 18
        summary_sheet.column_dimensions["B"].width = 28
        summary_sheet.column_dimensions["C"].width = 38
        summary_sheet.column_dimensions["D"].width = 14

    if "cadastros" in secoes:
        headers = [
            "ID", "Nome", "Nome social", "CPF", "RG", "Data do cadastro", "Status",
            "Data de nascimento", "Idade", "Faixa etária", "Sexo biológico",
            "Identidade de gênero", "Identidade étnico-racial", "Experiência de trabalho",
            "Tipo de ocupação", "Status ocupacional", "Grau de instrução", "Série final",
            "Deficiências", "Possui comorbidade", "Comorbidades/problemas de saúde",
            "Uso de substâncias psicoativas", "Substâncias informadas", "Cidade", "UF",
            "Bairro", "Endereço", "Telefone", "E-mail",
        ]
        rows = [
            [
                c.pk, c.nome, c.nome_social, c.cpf_numero, c.rg_numero, c.data_cadastro,
                c.get_status_display(), c.data_nascimento, _age_on(c.data_nascimento),
                _age_band_label(c.data_nascimento),
                c.get_sexo_biologico_display() if c.sexo_biologico else "",
                c.get_identidade_genero_display() if c.identidade_genero else "",
                c.get_identidade_etnico_racial_display() if c.identidade_etnico_racial else "",
                c.get_experiencia_trabalho_display() if c.experiencia_trabalho else "",
                c.get_tipo_ocupacao_display() if c.tipo_ocupacao else "",
                c.get_status_ocupacional_display() if c.status_ocupacional else "",
                c.get_grau_instrucao_display() if c.grau_instrucao else "",
                c.serie_concluida, c.get_deficiencias_display(),
                c.get_possui_comorbidade_display() if c.possui_comorbidade else "",
                c.comorbidades,
                c.get_uso_substancias_psicoativas_display() if c.uso_substancias_psicoativas else "",
                c.substancias_psicoativas, c.cidade, c.estado_uf, c.bairro, c.endereco,
                c.telefone_numero, c.email_contato,
            ]
            for c in cadastros
        ]
        add_sheet("Egressos", f"ESPI — Cadastros / Egressos ({len(rows)})", headers, rows)

    if "familiares" in secoes:
        headers = [
            "ID", "Nome", "Nome social", "CPF", "Data do cadastro", "Egresso vinculado",
            "Interno de referência", "Parentesco", "Data de nascimento", "Idade",
            "Faixa etária", "Sexo biológico", "Identidade de gênero",
            "Identidade étnico-racial", "Experiência de trabalho", "Ocupação/profissão",
            "Tipo de ocupação", "Grau de escolaridade", "Série final", "Deficiências",
            "Possui comorbidade", "Comorbidades/problemas de saúde",
            "Uso de substâncias psicoativas", "Substâncias informadas", "Bairro",
            "Telefone", "Contato", "E-mail",
        ]
        rows = [
            [
                f.pk, f.nome, f.nome_social, f.cpf_numero, f.data_criacao.date(),
                f.cadastro.nome if f.cadastro else "", f.nome_interno_referencia,
                f.parentesco, f.data_nascimento, _age_on(f.data_nascimento),
                _age_band_label(f.data_nascimento),
                f.get_sexo_biologico_display() if f.sexo_biologico else "",
                f.get_identidade_genero_display() if f.identidade_genero else "",
                f.get_identidade_etnico_racial_display() if f.identidade_etnico_racial else "",
                f.get_experiencia_trabalho_display() if f.experiencia_trabalho else "",
                f.ocupacao, f.get_tipo_ocupacao_display() if f.tipo_ocupacao else "",
                f.get_grau_instrucao_display() if f.grau_instrucao else "",
                f.serie_concluida, f.get_deficiencias_display(),
                f.get_possui_comorbidade_display() if f.possui_comorbidade else "",
                f.comorbidades,
                f.get_uso_substancias_psicoativas_display() if f.uso_substancias_psicoativas else "",
                f.substancias_psicoativas, f.bairro, f.telefone_numero,
                f.telefone_contato, f.email_contato,
            ]
            for f in familiares
        ]
        add_sheet("Familiares", f"ESPI — Familiares ({len(rows)})", headers, rows, "8E44AD")

    if "atendimentos" in secoes:
        headers = [
            "ID", "Data", "Pessoa atendida", "Perfil", "Tipo", "Local", "Motivo",
            "Objetivo", "Profissional responsável", "Outros participantes",
            "Descrição", "Status",
        ]
        rows = [
            [
                a.pk, a.data_atendimento, a.nome_pessoa_atendida,
                a.get_perfil_pessoa_atendida_display() if a.perfil_pessoa_atendida else "",
                a.get_tipo_atendimento_display() if a.tipo_atendimento else "",
                a.get_local_atendimento_display() if a.local_atendimento else "",
                a.get_motivo_procura_display() if a.motivo_procura else "",
                a.objetivo_atendimento, a.profissional_responsavel,
                a.outras_pessoas_participantes, a.descricao_atendimento,
                a.get_status_display() if a.status else "",
            ]
            for a in atendimentos
        ]
        add_sheet("Atendimentos", f"ESPI — Atendimentos ({len(rows)})", headers, rows, "00A86B")

    if "agendamentos" in secoes:
        headers = ["ID", "Data", "Horário", "Nome", "Tipo", "Observações"]
        rows = [
            [
                a.pk, a.data_agendamento, a.horario_atendimento, a.nome_atendido,
                a.get_tipo_agendamento_display() if a.tipo_agendamento else "",
                a.observacoes,
            ]
            for a in agendamentos
        ]
        add_sheet("Agendamentos", f"ESPI — Agendamentos ({len(rows)})", headers, rows, "F28C28")

    if not workbook.sheetnames:
        add_sheet("Relatório", "ESPI — Relatório", ["Informação"], [["Nenhuma seção selecionada"]])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    now = timezone.localtime(timezone.now())
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f"relatorio_espi_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
